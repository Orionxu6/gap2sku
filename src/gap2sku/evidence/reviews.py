from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from ..schemas.evidence import EvidenceRecord, RightsStatus, SourceType

try:
    import openpyxl  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover - checked by bootstrap
    openpyxl = None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value or "").strip()


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _looks_like_sku(value: str) -> bool:
    tokens = ("款", "灰", "绿", "紫", "凝胶", "收纳袋", "耳塞", "常规", "竹炭")
    return len(value) < 90 and any(token in value for token in tokens)


@dataclass
class ImportResult:
    records: list[EvidenceRecord]
    report: dict[str, Any]


class ReviewWorkbookImporter:
    """Imports review rows without changing the source workbook."""

    EXPECTED_COUNTS = {"尼拉": 69, "睡洞": 103, "祺加": 122, "西诺思": 95}

    def __init__(self, source_dir: str | Path) -> None:
        self.source_dir = Path(source_dir)

    def import_all(self) -> ImportResult:
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to import XLSX evidence")
        records: list[EvidenceRecord] = []
        workbook_stats: dict[str, dict[str, Any]] = {}
        for path in sorted(self.source_dir.rglob("*.xlsx")):
            brand = next((name for name in self.EXPECTED_COUNTS if name in path.name), path.stem)
            imported, stats = self._import_workbook(path, brand)
            records.extend(imported)
            workbook_stats[brand] = stats
        self._mark_duplicates(records)
        distribution = Counter(record.metadata["brand"] for record in records)
        if dict(distribution) != self.EXPECTED_COUNTS:
            raise ValueError(f"review distribution mismatch: expected {self.EXPECTED_COUNTS}, got {dict(distribution)}")
        report = {
            "total": len(records),
            "expected_total": 389,
            "distribution": dict(sorted(distribution.items())),
            "by_sentiment": dict(sorted(Counter(str(r.metadata.get("sentiment")) for r in records).items())),
            "source_url_count": sum(bool(r.source_url) for r in records),
            "source_label_count": sum(bool(r.metadata.get("source_label")) for r in records),
            "id_count": sum(bool(r.source_id) for r in records),
            "duplicate_count": sum(r.duplicate_of is not None for r in records),
            "column_migrations": sum("moved_sku_comment_columns" in r.transformations for r in records),
            "workbooks": workbook_stats,
            "limitations": [
                "定向截图/人工采样，不代表商品总体好评率或总体市场分布",
                "竞品评论不能证明供应商制造能力、BOM、合规或安全性",
                "缺少链接或评论 ID 的记录已降级，不得单独支撑 GO",
            ],
        }
        if report["total"] != 389:
            raise ValueError(f"expected 389 reviews, got {report['total']}")
        return ImportResult(records=records, report=report)

    def _import_workbook(self, path: Path, brand: str) -> tuple[list[EvidenceRecord], dict[str, Any]]:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["评论数据"]
        header_row = self._find_header(sheet)
        headers = [_text(cell.value) for cell in sheet[header_row]]
        file_hash = sha256_file(path)
        records: list[EvidenceRecord] = []
        for row_number, cells in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            values = list(cells)
            if len(values) < 7:
                continue
            review_id, product, observed, sentiment, rating, comment, sku = values[:7]
            comment_text, sku_text = _text(comment), _text(sku)
            transformations: list[str] = []
            # Rows XNS-R001..XNS-R065 have the two columns reversed in the
            # source workbook. Later rows use the normal layout. Use the row's
            # stable ID/product fields, not a fuzzy text-only guess.
            if brand == "西诺思" and (_text(review_id).startswith("XNS-") or "西诺思" in _text(product)):
                comment_text, sku_text = sku_text, comment_text
                transformations.append("moved_sku_comment_columns")
            sentiment_text = _text(sentiment)
            if not comment_text or sentiment_text not in {"好评", "中评", "差评"}:
                continue
            raw_source = _text(values[10] if brand == "睡洞" and len(values) > 10 else values[11] if len(values) > 11 else "")
            source_url = raw_source if raw_source.startswith(("http://", "https://")) else ""
            source_id = _text(review_id)
            product_text = _text(product)
            grade = "A" if source_id and product_text and source_url else "B" if raw_source else "C"
            content_hash = hashlib.sha256(comment_text.encode("utf-8")).hexdigest()
            evidence_id = f"ev-{brand}-{row_number:04d}-{content_hash[:10]}"
            metadata: dict[str, str | int | float | bool | None] = {
                "brand": brand, "product": product_text, "sentiment": sentiment_text,
                "rating": _text(rating), "sku": sku_text,
                "source_label": raw_source if not source_url else "",
                "follow_up": _text(values[7] if len(values) > 7 else ""),
                "has_image": _text(values[8] if len(values) > 8 else ""),
                "has_video": _text(values[9] if len(values) > 9 else ""),
            }
            records.append(EvidenceRecord(
                evidence_id=evidence_id, source_type=SourceType.USER_UPLOAD,
                source_id=source_id, source_url=source_url,
                snapshot_id=f"{path.stem}-sha256-{file_hash[:16]}", observed_at=_date(observed),
                content_excerpt=comment_text, content_hash=f"sha256:{content_hash}",
                locator=f"{path.name}#评论数据!R{row_number}", rights_status=RightsStatus.USER_AUTHORIZED,
                is_synthetic=False, language="zh-CN", file_hash=f"sha256:{file_hash}",
                workbook=path.name, sheet="评论数据", row_number=row_number,
                collected_at="2026-08-11", evidence_grade=grade,
                transformations=transformations, metadata=metadata,
            ))
        workbook.close()
        return records, {
            "file": path.name, "sha256": file_hash, "sheet": "评论数据",
            "header_row": header_row, "headers": headers, "imported": len(records),
        }

    @staticmethod
    def _find_header(sheet: Any) -> int:
        for row_number, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
            values = "|".join(_text(value) for value in row)
            if "评论内容" in values and ("评论ID" in values or "评价类型" in values):
                return row_number
        raise ValueError(f"cannot locate 评论数据 header in sheet {sheet.title}")

    @staticmethod
    def _mark_duplicates(records: list[EvidenceRecord]) -> None:
        seen: dict[tuple[str, str], str] = {}
        for record in records:
            key = (str(record.metadata.get("brand")), record.content_hash)
            if key in seen:
                record.duplicate_of = seen[key]
            else:
                seen[key] = record.evidence_id


def write_import_result(result: ImportResult, output_dir: str | Path) -> None:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    jsonl = output / "reviews.normalized.private.jsonl"
    jsonl.write_text(
        "".join(json.dumps(r.model_dump(mode="json"), ensure_ascii=False, sort_keys=True) + "\n" for r in result.records),
        encoding="utf-8",
    )
    (output / "import-report.json").write_text(
        json.dumps(result.report, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
